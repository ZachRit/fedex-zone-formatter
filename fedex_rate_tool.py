#!/usr/bin/env python3
"""
FedEx Rate Sheet Tool - Unified CLI for FedEx rate sheet operations.

Commands:
    find-pdfs      Find valid PDF URLs for postal codes
    parse-us-zones Parse US FedEx zone PDFs
    parse-ca-rates Parse Canadian FedEx rate PDFs
    generate       Generate rate sheets from zone data
    fix            Clean and deduplicate rate sheets

Usage:
    python fedex_rate_tool.py find-pdfs --input postal_codes.xlsx --output urls.txt
    python fedex_rate_tool.py parse-us-zones --input inputs/ --output outputs/
    python fedex_rate_tool.py parse-ca-rates --input CA_2026.pdf --origin M5V --output rates.xlsx
    python fedex_rate_tool.py generate --ssl-file ssls.xlsx --template template.xlsx --output outputs/
    python fedex_rate_tool.py fix --input outputs/ --output outputs/cleaned/
"""

import argparse
import re
import shutil
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
import pdfplumber
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font


# =============================================================================
# PDF URL Finder (from find_ranges_smart.py)
# =============================================================================

def check_url(url, timeout=8):
    """Check if a URL exists."""
    try:
        response = requests.head(url, timeout=timeout, allow_redirects=False,
                                 headers={'User-Agent': 'Mozilla/5.0'})
        return response.status_code == 200
    except Exception:
        return False


def find_range_containing(postal_code):
    """Find the PDF range that contains a given postal code."""
    pc = int(postal_code)

    # Try different starting points near the postal code
    possible_lowers = []

    # Try multiples of 100 below the postal code
    base = (pc // 100) * 100
    for offset in range(0, 2000, 100):
        lower = base - offset
        if lower >= 0:
            possible_lowers.append(lower)

    # For each possible lower, try different upper bounds
    sizes = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]

    for lower in possible_lowers:
        for size in sizes:
            upper = lower + size - 1
            if lower <= pc <= upper:
                url = f'https://www.fedex.com/ratetools/documents2/{lower:05d}-{upper:05d}.pdf'
                if check_url(url):
                    return (lower, upper, url)

    return None


def cmd_find_pdfs(args):
    """Find valid PDF URLs for postal codes."""
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1

    # Load postal codes
    df = pd.read_excel(input_path)

    # Try common column names
    postal_col = None
    for col in ['Postal Codes', 'Postal Code', 'postal_code', 'PostalCode', 'ZIP', 'zip']:
        if col in df.columns:
            postal_col = col
            break

    if postal_col is None:
        print(f"Error: Could not find postal code column. Available columns: {list(df.columns)}")
        return 1

    postal_codes = sorted(set([int(str(pc).zfill(5)) for pc in df[postal_col].tolist()]))

    print(f"Need to find ranges for {len(postal_codes)} postal codes")
    print(f"Postal codes: {[f'{pc:05d}' for pc in postal_codes[:10]]}... to {postal_codes[-1]:05d}")
    print("=" * 60)

    found_ranges = {}
    postal_to_range = {}

    for i, pc in enumerate(postal_codes):
        print(f"[{i+1}/{len(postal_codes)}] Finding range for {pc:05d}...", end='', flush=True)

        # Check if already covered
        already_covered = False
        for lower, (upper, url) in found_ranges.items():
            if lower <= pc <= upper:
                postal_to_range[pc] = (lower, upper)
                print(f" already covered by {lower:05d}-{upper:05d}")
                already_covered = True
                break

        if already_covered:
            continue

        result = find_range_containing(pc)
        if result:
            lower, upper, url = result
            found_ranges[lower] = (upper, url)
            postal_to_range[pc] = (lower, upper)
            print(f" FOUND: {lower:05d}-{upper:05d}")
        else:
            print(f" NOT FOUND!")

    print("\n" + "=" * 60)
    print(f"Found {len(found_ranges)} unique PDF ranges:")
    print("=" * 60)

    sorted_ranges = sorted(found_ranges.items())
    urls = [url for lower, (upper, url) in sorted_ranges]

    for url in urls:
        print(url)

    # Save to file
    output_path = Path(args.output)
    with open(output_path, 'w') as f:
        for url in urls:
            f.write(url + '\n')

    print(f"\nSaved {len(urls)} URLs to {output_path}")

    uncovered = [pc for pc in postal_codes if pc not in postal_to_range]
    if uncovered:
        print(f"\nWARNING: {len(uncovered)} postal codes not covered:")
        for pc in uncovered:
            print(f"  {pc:05d}")
    else:
        print(f"\nAll {len(postal_codes)} postal codes are covered!")

    return 0


# =============================================================================
# US Zone PDF Parser (from parse_fedex_zones.py)
# =============================================================================

def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract all text from PDF using pdfplumber."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text


def parse_contiguous_us(text: str) -> list:
    """Parse the Contiguous U.S. section."""
    results = []
    pattern = r'(\d{5}(?:-\d{5})?)\s+(\d+|NA|\*)'
    contiguous_match = re.search(r'Contiguous U\.S\.(.*?)Alaska,\s*Hawaii', text, re.DOTALL | re.IGNORECASE)

    if contiguous_match:
        contiguous_text = contiguous_match.group(1)
        matches = re.findall(pattern, contiguous_text)
        for zip_range, zone in matches:
            results.append((zip_range, zone))

    return results


def parse_alaska_hawaii_pr(text: str) -> list:
    """Parse the Alaska, Hawaii, and Puerto Rico section."""
    results = []
    pattern = r'(\d{5}(?:-\d{5})?)\s+(\d+|NA|\*)\s+(\d+|NA|\*)'
    alaska_match = re.search(r'Alaska,\s*Hawaii,?\s*and\s*Puerto\s*Rico(.*?)$', text, re.DOTALL | re.IGNORECASE)

    if alaska_match:
        alaska_text = alaska_match.group(1)
        matches = re.findall(pattern, alaska_text)
        for zip_range, express_zone, ground_zone in matches:
            results.append((zip_range, express_zone))

    return results


def split_zip_range(zip_range: str) -> tuple:
    """Split ZIP range into start and end."""
    if '-' in zip_range:
        start, end = zip_range.split('-')
        return (start.strip(), end.strip())
    else:
        return (zip_range.strip(), zip_range.strip())


def normalize_zone(zone: str) -> str:
    """Normalize zone value."""
    if zone.upper() == 'NA' or zone == '*':
        return ''
    return zone


def validate_fedex_pdf(text: str) -> bool:
    """Validate that the PDF appears to be a FedEx zone locator."""
    has_fedex = 'fedex' in text.lower()
    has_zone = 'zone' in text.lower()
    zip_pattern = r'\d{5}(?:-\d{5})?\s+\d+'
    has_zip_data = bool(re.search(zip_pattern, text))
    return has_fedex and has_zone and has_zip_data


def process_us_zone_pdf(input_path: str, output_dir: str = 'outputs') -> str:
    """Process a FedEx zone PDF and output to xlsx."""
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    text = extract_text_from_pdf(str(input_path))

    if not validate_fedex_pdf(text):
        raise ValueError("PDF does not appear to be a valid FedEx zone locator")

    contiguous_data = parse_contiguous_us(text)
    alaska_data = parse_alaska_hawaii_pr(text)

    if not contiguous_data and not alaska_data:
        raise ValueError("No zone data could be extracted from PDF")

    all_data = contiguous_data + alaska_data

    rows = []
    for zip_range, zone in all_data:
        start_zip, end_zip = split_zip_range(zip_range)
        normalized_zone = normalize_zone(zone)
        rows.append({
            'Start Postal Code': start_zip,
            'End Postal Code': end_zip,
            'Zone': normalized_zone
        })

    df = pd.DataFrame(rows)
    df = df.sort_values('Start Postal Code').reset_index(drop=True)

    output_filename = input_path.stem + '.xlsx'
    output_path = output_dir / output_filename

    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Zones')
        worksheet = writer.sheets['Zones']
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=2):
            for cell in row:
                cell.number_format = '@'

    return str(output_path)


def cmd_parse_us_zones(args):
    """Parse US FedEx zone PDFs."""
    input_path = Path(args.input)
    output_dir = Path(args.output)

    if input_path.is_file():
        # Single file
        print(f"Processing: {input_path}")
        try:
            output_path = process_us_zone_pdf(str(input_path), str(output_dir))
            print(f"Output written to: {output_path}")
            return 0
        except Exception as e:
            print(f"Error: {e}")
            return 1
    elif input_path.is_dir():
        # Directory of PDFs
        pdf_files = list(input_path.glob('*.pdf'))
        archive_dir = input_path / 'archive'
        failed_dir = input_path / 'failed_parsing'

        archive_dir.mkdir(parents=True, exist_ok=True)
        failed_dir.mkdir(parents=True, exist_ok=True)

        if not pdf_files:
            print(f"No PDF files found in {input_path}")
            return 0

        print(f"Found {len(pdf_files)} PDF file(s) to process")
        success_count = 0
        failed_count = 0

        for pdf_path in pdf_files:
            print(f"Processing: {pdf_path.name}")
            try:
                output_path = process_us_zone_pdf(str(pdf_path), str(output_dir))
                print(f"  Output: {output_path}")
                shutil.move(str(pdf_path), str(archive_dir / pdf_path.name))
                print(f"  Moved to: {archive_dir / pdf_path.name}")
                success_count += 1
            except Exception as e:
                print(f"  ERROR: {e}")
                shutil.move(str(pdf_path), str(failed_dir / pdf_path.name))
                print(f"  Moved to: {failed_dir / pdf_path.name}")
                failed_count += 1

        print("=" * 40)
        print(f"Complete: {success_count} succeeded, {failed_count} failed")
        return 0 if failed_count == 0 else 1
    else:
        print(f"Error: {input_path} is not a file or directory")
        return 1


# =============================================================================
# Canadian Rate PDF Parser (from parse_ca_fedex_rate_sheets.py)
# =============================================================================

ZONE_CODES = [
    'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ',
    'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT',
    'DU', 'DV', 'DW', 'DX', 'DY', 'DZ'
]

SERVICE_DEFINITIONS = [
    {"name": "FedEx First Overnight", "search": "FedEx First Overnight", "page_count": 4, "is_freight": False},
    {"name": "FedEx Priority Overnight", "search": "FedEx Priority Overnight", "page_count": 4, "is_freight": False},
    {"name": "FedEx Standard Overnight", "search": "FedEx Standard Overnight", "page_count": 4, "is_freight": False},
    {"name": "FedEx 2Day", "search": "FedEx 2Day", "page_count": 4, "is_freight": False},
    {"name": "FedEx Economy", "search": "FedEx Economy", "page_count": 4, "is_freight": False},
    {"name": "FedEx 1Day Freight", "search": "FedEx 1Day", "page_count": 2, "is_freight": True},
]


def detect_service_pages(pdf):
    """Dynamically detect page ranges for each service."""
    services = []

    for service_def in SERVICE_DEFINITIONS:
        search_term = service_def["search"]
        start_page = None

        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            lines = text.split('\n')

            title_found = False
            for line in lines[:5]:
                if search_term in line and "Rates" in line:
                    title_found = True
                    break

            if title_found:
                already_assigned = False
                for s in services:
                    if s["pages"][0] <= i <= s["pages"][1]:
                        already_assigned = True
                        break

                if not already_assigned:
                    start_page = i
                    break

        if start_page is not None:
            end_page = start_page + service_def["page_count"] - 1
            services.append({
                "name": service_def["name"],
                "pages": (start_page, end_page),
                "is_freight": service_def["is_freight"]
            })

    return services


def find_zone_index_pages(pdf):
    """Find the pages containing the Zone Index tables."""
    postal_code_page = None
    zone_matrix_page = None

    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""

        if "Postal Code Zone Index" in text and postal_code_page is None:
            postal_code_page = i

        if "Intra-Canada Zone Index" in text and zone_matrix_page is None:
            zone_matrix_page = i

    return postal_code_page, zone_matrix_page


def parse_postal_code_to_zone_mapping(pdf, postal_code_page):
    """Parse the Postal Code Zone Index page."""
    if postal_code_page is None:
        return {}

    page = pdf.pages[postal_code_page]
    text = page.extract_text() or ""

    postal_zone_map = {}

    range_pattern = re.compile(r'([A-Z]\d[A-Z])\s*[—–-]\s*([A-Z]\d[A-Z])\s+(D[A-Z])')

    for match in range_pattern.finditer(text):
        start_postal = match.group(1)
        end_postal = match.group(2)
        zone_code = match.group(3)
        postal_zone_map[(start_postal, end_postal)] = zone_code

    lines = text.split('\n')
    for line in lines:
        if 'Postal Code' in line and 'Zone' in line:
            continue

        parts = line.split()
        i = 0
        while i < len(parts):
            part = parts[i]
            if re.match(r'^[A-Z]\d[A-Z]$', part):
                if i + 1 < len(parts) and re.match(r'^D[A-Z]$', parts[i + 1]):
                    zone_code = parts[i + 1]
                    if i > 0 and parts[i - 1] in ['—', '–', '-']:
                        i += 2
                        continue
                    already_covered = False
                    for (start, end), _ in postal_zone_map.items():
                        if start == part and end != part:
                            already_covered = True
                            break
                    if not already_covered:
                        postal_zone_map[(part, part)] = zone_code
            i += 1

    return postal_zone_map


def parse_zone_matrix(pdf, zone_matrix_page):
    """Parse the Intra-Canada Zone Index matrix."""
    if zone_matrix_page is None:
        return {}

    page = pdf.pages[zone_matrix_page]
    tables = page.extract_tables()

    if not tables:
        return {}

    zone_table = max(tables, key=lambda t: len(t))
    zone_matrix = {}
    dest_zones = ZONE_CODES.copy()

    for row_idx, row in enumerate(zone_table):
        if not row or len(row) < 2:
            continue

        first_cell = str(row[0]) if row[0] else ''
        second_cell = str(row[1]) if row[1] else ''

        if 'Origin' in first_cell or 'Destination' in first_cell:
            continue
        if 'DA DB DC' in second_cell:
            continue

        origin_cell = first_cell.strip()
        values_cell = second_cell.strip()

        if not origin_cell or not values_cell:
            continue

        origin_lines = origin_cell.split('\n')
        values_lines = values_cell.split('\n')

        for i, origin_line in enumerate(origin_lines):
            origin_zone = origin_line.strip()
            if origin_zone not in ZONE_CODES:
                continue

            if i < len(values_lines):
                values_str = values_lines[i].strip()
            else:
                continue

            values = []
            for val in values_str.split():
                try:
                    values.append(int(val))
                except ValueError:
                    continue

            if len(values) == len(dest_zones):
                for j, dest_zone in enumerate(dest_zones):
                    zone_matrix[(origin_zone, dest_zone)] = values[j]

    return zone_matrix


def get_zone_code_for_postal(postal_code, postal_zone_map):
    """Find which zone code a postal code maps to."""
    fsa = postal_code[:3].upper()

    for (start_postal, end_postal), zone_code in postal_zone_map.items():
        if start_postal <= fsa <= end_postal:
            return zone_code

    return None


def generate_zones_data(origin_postal, postal_zone_map, zone_matrix):
    """Generate zone data for all destination postal codes."""
    origin_zone = get_zone_code_for_postal(origin_postal, postal_zone_map)

    if origin_zone is None:
        print(f"Warning: Could not find zone code for origin postal code: {origin_postal}")
        return []

    print(f"  Origin postal code {origin_postal} maps to zone {origin_zone}")

    zones_data = []

    for (start_postal, end_postal), dest_zone_code in sorted(postal_zone_map.items()):
        matrix_key = (origin_zone, dest_zone_code)
        if matrix_key in zone_matrix:
            numerical_zone = zone_matrix[matrix_key]
        else:
            numerical_zone = 16

        zones_data.append((
            "Canada",
            "CA",
            numerical_zone,
            "",
            start_postal,
            end_postal
        ))

    return zones_data


def clean_rate(value):
    """Clean a rate value string and convert to Decimal."""
    if value is None:
        return None
    cleaned = str(value).replace('$', '').replace(',', '').replace(' ', '').strip()
    if not cleaned or cleaned == '—' or cleaned == '-':
        return None
    try:
        return Decimal(cleaned)
    except Exception:
        return None


def parse_weight_from_line(line):
    """Parse weight from a line."""
    line = line.strip()
    if not line:
        return None

    match = re.match(r'(\d+)\s*lbs?\.?', line)
    if match:
        return int(match.group(1))

    parts = line.split()
    if len(parts) >= 1:
        try:
            weight = int(parts[0])
            if 1 <= weight <= 2000:
                return weight
        except ValueError:
            pass

    return None


def parse_zone_numbers(header_text):
    """Parse zone numbers from a header cell."""
    if not header_text:
        return []
    parts = str(header_text).split()
    zones = []
    for part in parts:
        try:
            zone = int(part)
            if 1 <= zone <= 16:
                zones.append(zone)
        except ValueError:
            continue
    return zones


def parse_rates_line(rates_text):
    """Parse rates from a line."""
    if not rates_text:
        return []
    parts = str(rates_text).replace('$', '').split()
    rates = []
    for part in parts:
        rate = clean_rate(part)
        if rate is not None:
            rates.append(rate)
    return rates


def parse_non_freight_rates(pdf, start_page, end_page):
    """Parse rates for non-freight services."""
    rates = {}
    per_pound_rates = {}

    pages = [pdf.pages[i] for i in range(start_page, end_page + 1)]

    for page_idx, page in enumerate(pages):
        all_tables = page.extract_tables()
        if not all_tables:
            continue

        for raw_table in all_tables:
            table = []
            for row in raw_table:
                if not row:
                    continue
                max_lines = 1
                cell_lines = []
                for cell in row:
                    if cell is None:
                        cell_lines.append([''])
                    else:
                        lines = str(cell).split('\n')
                        cell_lines.append(lines)
                        max_lines = max(max_lines, len(lines))

                if max_lines == 1:
                    table.append(row)
                else:
                    for line_idx in range(max_lines):
                        new_row = []
                        for lines in cell_lines:
                            if line_idx < len(lines):
                                new_row.append(lines[line_idx].strip())
                            else:
                                new_row.append('')
                        table.append(new_row)

            if not table:
                continue

            zones = []

            for row in table[:5]:
                for cell in row:
                    if cell:
                        found_zones = parse_zone_numbers(cell)
                        if len(found_zones) >= 7:
                            zones = found_zones
                            break
                if zones:
                    break

            if not zones:
                continue

            for row in table:
                if not row or len(row) < 2:
                    continue

                weight_cell = str(row[0]).strip() if row[0] else ''
                rates_cell = str(row[1]).strip() if len(row) > 1 and row[1] else ''

                if 'weight' in weight_cell.lower() or 'zone' in weight_cell.lower():
                    continue
                if 'envelope' in weight_cell.lower() or 'pak' in weight_cell.lower():
                    continue

                if '100 lbs' in weight_cell.lower() or '100lbs' in weight_cell.lower().replace(' ', ''):
                    rate_values = parse_rates_line(rates_cell)
                    if len(rate_values) == len(zones):
                        for i, zone in enumerate(zones):
                            per_pound_rates[zone] = rate_values[i]
                    continue

                weight = parse_weight_from_line(weight_cell)
                if weight is None or weight < 1 or weight > 99:
                    continue

                rate_values = parse_rates_line(rates_cell)

                if not rate_values:
                    continue

                if weight not in rates:
                    rates[weight] = {}

                for i, rate in enumerate(rate_values):
                    if i < len(zones):
                        rates[weight][zones[i]] = rate

    for weight in range(100, 151):
        rates[weight] = {}
        for zone in range(1, 17):
            if zone in per_pound_rates:
                rate = (Decimal(weight) * per_pound_rates[zone]).quantize(
                    Decimal('0.01'), rounding=ROUND_HALF_UP
                )
                rates[weight][zone] = rate

    return rates


def parse_freight_rates(pdf, start_page, end_page):
    """Parse rates for freight services."""
    rates = {}
    per_pound_rates = {}
    minimum_charges = {}

    pages = [pdf.pages[i] for i in range(start_page, end_page + 1)]

    brackets = {
        "151 to 299": (151, 299),
        "300 to 499": (300, 499),
        "500 to 999": (500, 999),
        "1000 to 1999": (1000, 1999),
        "1000 to1999": (1000, 1999),
        "2000 or more": (2000, 2000),
    }

    for page_idx, page in enumerate(pages):
        tables = page.extract_tables()
        if not tables:
            continue

        raw_table = tables[0]

        table = []
        for row in raw_table:
            if not row:
                continue
            max_lines = 1
            cell_lines = []
            for cell in row:
                if cell is None:
                    cell_lines.append([''])
                else:
                    lines = str(cell).split('\n')
                    cell_lines.append(lines)
                    max_lines = max(max_lines, len(lines))

            if max_lines == 1:
                table.append(row)
            else:
                for line_idx in range(max_lines):
                    new_row = []
                    for lines in cell_lines:
                        if line_idx < len(lines):
                            new_row.append(lines[line_idx].strip())
                        else:
                            new_row.append('')
                    table.append(new_row)

        zones = []

        for row in table:
            if not row or len(row) < 2:
                continue

            weight_cell = str(row[0]).strip() if row[0] else ''
            rates_cell = str(row[1]).strip() if len(row) > 1 and row[1] else ''

            if not zones:
                found_zones = parse_zone_numbers(rates_cell)
                if len(found_zones) >= 7:
                    zones = found_zones
                    continue

            if not zones:
                continue

            if 'minimum' in weight_cell.lower():
                rate_values = parse_rates_line(rates_cell)
                if len(rate_values) == len(zones):
                    for i, zone in enumerate(zones):
                        minimum_charges[zone] = rate_values[i]
                continue

            for bracket_name, (bracket_start, bracket_end) in brackets.items():
                if bracket_name.lower().replace(' ', '') in weight_cell.lower().replace(' ', ''):
                    rate_values = parse_rates_line(rates_cell)
                    if len(rate_values) == len(zones):
                        bracket_key = (bracket_start, bracket_end)
                        if bracket_key not in per_pound_rates:
                            per_pound_rates[bracket_key] = {}
                        for i, zone in enumerate(zones):
                            per_pound_rates[bracket_key][zone] = rate_values[i]
                    break

    for weight in range(151, 2001):
        rates[weight] = {}

        if weight <= 299:
            bracket_key = (151, 299)
        elif weight <= 499:
            bracket_key = (300, 499)
        elif weight <= 999:
            bracket_key = (500, 999)
        elif weight <= 1999:
            bracket_key = (1000, 1999)
        else:
            bracket_key = (2000, 2000)

        if bracket_key not in per_pound_rates:
            continue

        for zone in range(1, 17):
            if zone not in per_pound_rates[bracket_key]:
                continue

            per_lb_rate = per_pound_rates[bracket_key][zone]
            calculated_rate = (Decimal(weight) * per_lb_rate).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            )

            if zone in minimum_charges:
                calculated_rate = max(calculated_rate, minimum_charges[zone])

            rates[weight][zone] = calculated_rate

    return rates


def write_ca_excel(services_data, output_path, zones_data=None):
    """Write Canadian rate data to Excel file."""
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    if zones_data:
        ws_zones = wb.create_sheet(title="Zones", index=0)

        ws_zones.cell(row=1, column=1, value="Destination Zones")
        ws_zones.cell(row=1, column=1).font = Font(bold=True, size=14)

        ws_zones.cell(row=2, column=1, value="Zones should be a number from 1 to 16")

        headers = ["Country Name", "Country Symbol", "Zone", "City", "Start Postal Code", "End Postal Code"]
        for col, header in enumerate(headers, start=1):
            ws_zones.cell(row=3, column=col, value=header)
            ws_zones.cell(row=3, column=col).font = Font(bold=True)

        for row_num, (country_name, country_symbol, zone, city, start_postal, end_postal) in enumerate(zones_data, start=4):
            ws_zones.cell(row=row_num, column=1, value=country_name)
            ws_zones.cell(row=row_num, column=2, value=country_symbol)
            ws_zones.cell(row=row_num, column=3, value=zone)
            ws_zones.cell(row=row_num, column=4, value=city)
            ws_zones.cell(row=row_num, column=5, value=start_postal)
            ws_zones.cell(row=row_num, column=6, value=end_postal)

        ws_zones.column_dimensions['A'].width = 15
        ws_zones.column_dimensions['B'].width = 15
        ws_zones.column_dimensions['C'].width = 8
        ws_zones.column_dimensions['D'].width = 15
        ws_zones.column_dimensions['E'].width = 18
        ws_zones.column_dimensions['F'].width = 18

    for service_name, rates, is_freight in services_data:
        sheet_name = service_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value=service_name)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        ws.cell(row=2, column=1, value="Rates are specified in ($) CAD: 2025")

        ws.cell(row=3, column=1, value="Weight (lb)")
        ws.cell(row=3, column=1).font = Font(bold=True)
        for zone in range(1, 17):
            ws.cell(row=3, column=zone + 1, value=f"Zone {zone}")
            ws.cell(row=3, column=zone + 1).font = Font(bold=True)

        if is_freight:
            weight_range = range(151, 2001)
        else:
            weight_range = range(1, 151)

        row_num = 4
        for weight in weight_range:
            if weight not in rates:
                continue

            ws.cell(row=row_num, column=1, value=weight)

            for zone in range(1, 17):
                if zone in rates[weight]:
                    ws.cell(row=row_num, column=zone + 1, value=float(rates[weight][zone]))

            row_num += 1

        ws.column_dimensions['A'].width = 12
        for col in range(2, 18):
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 10

    wb.save(output_path)
    print(f"Saved to {output_path}")


def cmd_parse_ca_rates(args):
    """Parse Canadian FedEx rate PDFs."""
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1

    print(f"Parsing {input_path}...")

    services_data = []
    zones_data = None

    with pdfplumber.open(input_path) as pdf:
        if args.origin:
            print("\nParsing Zone Index...")

            postal_code_page, zone_matrix_page = find_zone_index_pages(pdf)

            if postal_code_page is not None:
                print(f"  Found Postal Code Zone Index on page {postal_code_page + 1}")
            else:
                print("  Warning: Could not find Postal Code Zone Index page")

            if zone_matrix_page is not None:
                print(f"  Found Intra-Canada Zone Index on page {zone_matrix_page + 1}")
            else:
                print("  Warning: Could not find Intra-Canada Zone Index page")

            postal_zone_map = parse_postal_code_to_zone_mapping(pdf, postal_code_page)
            print(f"  Parsed {len(postal_zone_map)} postal code ranges")

            zone_matrix = parse_zone_matrix(pdf, zone_matrix_page)
            print(f"  Parsed {len(zone_matrix)} zone matrix entries")

            zones_data = generate_zones_data(
                args.origin,
                postal_zone_map,
                zone_matrix
            )
            print(f"  Generated {len(zones_data)} destination zone entries")

        services = detect_service_pages(pdf)

        print("\nParsing Rate Sheets...")
        for service in services:
            print(f"  Processing {service['name']}...")

            start_page, end_page = service['pages']

            if service['is_freight']:
                rates = parse_freight_rates(pdf, start_page, end_page)
            else:
                rates = parse_non_freight_rates(pdf, start_page, end_page)

            total_rates = sum(len(zones) for zones in rates.values())
            print(f"    Extracted {len(rates)} weights, {total_rates} rate values")

            services_data.append((service['name'], rates, service['is_freight']))

    print(f"\nWriting to {args.output}...")
    write_ca_excel(services_data, args.output, zones_data)

    print("Done!")
    return 0


# =============================================================================
# Rate Sheet Generator (from generate_rate_sheet.py)
# =============================================================================

def find_zone_file(postal_code: str, outputs_dir: str = 'outputs') -> Path:
    """Find the zone xlsx file whose filename range contains the postal code."""
    outputs_path = Path(outputs_dir)

    postal_code = str(postal_code).zfill(5)
    postal_int = int(postal_code)

    for zone_file in outputs_path.glob('*.xlsx'):
        if zone_file.name.startswith('~$'):
            continue

        stem = zone_file.stem
        if '-' not in stem:
            continue

        try:
            start_str, end_str = stem.split('-')
            start_range = int(start_str)
            end_range = int(end_str)

            if start_range <= postal_int <= end_range:
                return zone_file
        except ValueError:
            continue

    return None


def load_zone_data(zone_file: Path, country_name: str, country_symbol: str) -> pd.DataFrame:
    """Load zone file and add Country Name, Country Symbol columns."""
    df = pd.read_excel(zone_file)

    df['Country Name'] = country_name
    df['Country Symbol'] = country_symbol
    df['City'] = ''

    df['Start Postal Code'] = df['Start Postal Code'].apply(lambda x: str(int(x)).zfill(5) if pd.notna(x) else '')
    df['End Postal Code'] = df['End Postal Code'].apply(lambda x: str(int(x)).zfill(5) if pd.notna(x) else '')

    df['Zones'] = df['Zone'].apply(lambda x: str(int(x)) if pd.notna(x) else '')

    df = df[['Country Name', 'Country Symbol', 'Zones', 'City', 'Start Postal Code', 'End Postal Code']]

    return df


def append_to_rate_sheet(rate_sheet_path: Path, zone_data: pd.DataFrame):
    """Load rate sheet, append zone data to Zone tab."""
    wb = load_workbook(rate_sheet_path)

    if 'Zones' not in wb.sheetnames:
        raise ValueError(f"Rate sheet does not contain a 'Zones' tab")

    ws = wb['Zones']

    start_row = 4

    for idx, row in zone_data.iterrows():
        ws.cell(row=start_row + idx, column=1, value=row['Country Name'])
        ws.cell(row=start_row + idx, column=2, value=row['Country Symbol'])
        ws.cell(row=start_row + idx, column=3, value=row['Zones'])
        ws.cell(row=start_row + idx, column=4, value=row['City'])
        ws.cell(row=start_row + idx, column=5, value=row['Start Postal Code'])
        ws.cell(row=start_row + idx, column=6, value=row['End Postal Code'])

    return wb


def generate_output_filename(ssl: str, client_name: str, carrier: str, carrier_account: str) -> str:
    """Generate output filename."""
    date_str = datetime.now().strftime('%Y%m%d')
    return f"{date_str}-{ssl}-{client_name}-{carrier}-{carrier_account}.xlsx"


def cmd_generate(args):
    """Generate rate sheets from zone data."""
    input_path = Path(args.ssl_file)
    rate_sheet_path = Path(args.template)
    output_dir = Path(args.output)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1

    if not rate_sheet_path.exists():
        print(f"Error: Rate sheet template not found: {rate_sheet_path}")
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        ssl_df = pd.read_excel(input_path)
    except Exception as e:
        print(f"Error reading input file: {e}")
        return 1

    required_cols = ['SSL', 'Postal Code']
    missing_cols = [col for col in required_cols if col not in ssl_df.columns]
    if missing_cols:
        print(f"Error: Input file missing required columns: {missing_cols}")
        return 1

    try:
        wb_test = load_workbook(rate_sheet_path)
        if 'Zones' not in wb_test.sheetnames:
            print(f"Error: Rate sheet does not contain a 'Zones' tab")
            return 1
        wb_test.close()
    except Exception as e:
        print(f"Error reading rate sheet: {e}")
        return 1

    grouped = ssl_df.groupby('SSL')

    print(f"Processing {len(grouped)} SSL group(s)...")

    for ssl, group in grouped:
        print(f"Processing SSL: {ssl}")

        all_zone_data = []

        for _, row in group.iterrows():
            postal_code = str(row['Postal Code']).zfill(5)

            zone_file = find_zone_file(postal_code, args.zones_dir)

            if zone_file is None:
                print(f"  Warning: No zone file found for postal code {postal_code}, skipping")
                continue

            print(f"  Found zone file for {postal_code}: {zone_file.name}")

            zone_data = load_zone_data(zone_file, args.country_name, args.country_symbol)
            all_zone_data.append(zone_data)

        if not all_zone_data:
            print(f"  Warning: No zone data collected for SSL {ssl}, skipping")
            continue

        combined_zone_data = pd.concat(all_zone_data, ignore_index=True)

        wb = append_to_rate_sheet(rate_sheet_path, combined_zone_data)

        output_filename = generate_output_filename(
            ssl, args.client_name, args.carrier, args.carrier_account
        )
        output_path = output_dir / output_filename
        wb.save(output_path)
        wb.close()

        print(f"  Output: {output_path} ({len(combined_zone_data)} rows)")

    print("Complete!")
    return 0


# =============================================================================
# Rate Sheet Fixer (from fix_rate_sheets.py)
# =============================================================================

def fix_zone_headers(ws) -> int:
    """Fix 'Zone 0X' -> 'Zone X' in row 3 headers."""
    fixes = 0
    pattern = re.compile(r'^Zone 0(\d)$')

    for cell in ws[3]:
        if isinstance(cell, MergedCell):
            continue
        if cell.value and isinstance(cell.value, str):
            match = pattern.match(cell.value)
            if match:
                cell.value = f"Zone {match.group(1)}"
                fixes += 1

    return fixes


def deduplicate_zones_tab(wb) -> int:
    """Deduplicate Zones tab."""
    if 'Zones' not in wb.sheetnames:
        return 0

    ws = wb['Zones']

    data = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if any(cell is not None for cell in row):
            data.append(row)

    if not data:
        return 0

    header_row = [cell.value for cell in ws[3]]
    num_cols = len([h for h in header_row if h is not None])

    df = pd.DataFrame(data)

    col_indices = {}
    for idx, header in enumerate(header_row):
        if header in ['Country Symbol', 'Zone', 'Zones', 'Start Postal Code', 'End Postal Code']:
            col_indices[header] = idx

    dedup_cols = []
    for name in ['Country Symbol', 'Zone', 'Zones', 'Start Postal Code', 'End Postal Code']:
        if name in col_indices:
            dedup_cols.append(col_indices[name])

    if not dedup_cols:
        return 0

    original_count = len(df)
    df_deduped = df.drop_duplicates(subset=dedup_cols, keep='first')
    removed_count = original_count - len(df_deduped)

    if removed_count == 0:
        return 0

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    for row_idx, row_data in enumerate(df_deduped.values, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value

    return removed_count


def process_fix_file(filepath: Path, output_dir: Path, processed_dir: Path) -> dict:
    """Process a single rate sheet file."""
    stats = {'header_fixes': 0, 'duplicates_removed': 0, 'success': False}

    try:
        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            if sheet_name != 'Zones':
                stats['header_fixes'] += fix_zone_headers(wb[sheet_name])

        stats['duplicates_removed'] = deduplicate_zones_tab(wb)

        output_path = output_dir / filepath.name
        wb.save(output_path)
        wb.close()

        shutil.move(str(filepath), str(processed_dir / filepath.name))

        stats['success'] = True

    except Exception as e:
        stats['error'] = str(e)

    return stats


def cmd_fix(args):
    """Clean and deduplicate rate sheets."""
    input_dir = Path(args.input)
    output_dir = Path(args.output)

    if not input_dir.exists():
        print(f"Error: Directory not found: {input_dir}")
        return 1

    processed_dir = input_dir / 'processed'

    output_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(exist_ok=True)

    xlsx_files = [f for f in input_dir.glob('*.xlsx') if not f.name.startswith('~$')]

    if not xlsx_files:
        print(f"No .xlsx files found in {input_dir}")
        return 0

    print(f"Processing {len(xlsx_files)} files...")

    total_header_fixes = 0
    total_duplicates_removed = 0
    success_count = 0

    for filepath in xlsx_files:
        print(f"Processing: {filepath.name}")

        stats = process_fix_file(filepath, output_dir, processed_dir)

        if stats['success']:
            success_count += 1
            total_header_fixes += stats['header_fixes']
            total_duplicates_removed += stats['duplicates_removed']
            print(f"  Header fixes: {stats['header_fixes']}, Duplicates removed: {stats['duplicates_removed']}")
        else:
            print(f"  Error: {stats.get('error', 'Unknown error')}")

    print()
    print(f"Complete! {success_count}/{len(xlsx_files)} files processed")
    print(f"Total header fixes: {total_header_fixes}")
    print(f"Total duplicates removed: {total_duplicates_removed}")
    print(f"Output: {output_dir}")
    print(f"Originals moved to: {processed_dir}")

    return 0


# =============================================================================
# Main CLI Entry Point
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='FedEx Rate Sheet Tool - Unified CLI for FedEx rate sheet operations',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python fedex_rate_tool.py find-pdfs --input inputs/postal_codes.xlsx --output urls.txt
    python fedex_rate_tool.py parse-us-zones --input inputs/ --output outputs/
    python fedex_rate_tool.py parse-ca-rates --input inputs/CA_2026.pdf --origin M5V --output outputs/rates.xlsx
    python fedex_rate_tool.py generate --ssl-file inputs/ssls.xlsx --template template.xlsx --output outputs/
    python fedex_rate_tool.py fix --input outputs/ --output outputs/cleaned/
        """
    )

    subparsers = parser.add_subparsers(dest='command', required=True, help='Command to run')

    # find-pdfs subcommand
    find_parser = subparsers.add_parser('find-pdfs', help='Find valid PDF URLs for postal codes')
    find_parser.add_argument('--input', '-i', required=True, help='Excel file with postal codes')
    find_parser.add_argument('--output', '-o', default='valid_pdf_urls.txt', help='Output file for URLs')

    # parse-us-zones subcommand
    parse_us_parser = subparsers.add_parser('parse-us-zones', help='Parse US FedEx zone PDFs')
    parse_us_parser.add_argument('--input', '-i', required=True, help='PDF file or directory of PDFs')
    parse_us_parser.add_argument('--output', '-o', default='outputs', help='Output directory')

    # parse-ca-rates subcommand
    parse_ca_parser = subparsers.add_parser('parse-ca-rates', help='Parse Canadian FedEx rate PDFs')
    parse_ca_parser.add_argument('--input', '-i', required=True, help='Path to Canadian rate PDF')
    parse_ca_parser.add_argument('--origin', '-p', help='Origin postal code (FSA) for zone calculations')
    parse_ca_parser.add_argument('--output', '-o', default='CA_Express_Rates.xlsx', help='Output Excel file')

    # generate subcommand
    gen_parser = subparsers.add_parser('generate', help='Generate rate sheets from zone data')
    gen_parser.add_argument('--ssl-file', required=True, help='Excel file with SSL and Postal Code columns')
    gen_parser.add_argument('--template', required=True, help='Rate sheet template Excel file')
    gen_parser.add_argument('--zones-dir', default='outputs', help='Directory containing zone files')
    gen_parser.add_argument('--country-name', default='United States', help='Country name')
    gen_parser.add_argument('--country-symbol', default='US', help='Country symbol')
    gen_parser.add_argument('--client-name', required=True, help='Client name for output filename')
    gen_parser.add_argument('--carrier', default='FedEx', help='Carrier name')
    gen_parser.add_argument('--carrier-account', required=True, help='Carrier account number')
    gen_parser.add_argument('--output', '-o', default='outputs', help='Output directory')

    # fix subcommand
    fix_parser = subparsers.add_parser('fix', help='Clean and deduplicate rate sheets')
    fix_parser.add_argument('--input', '-i', required=True, help='Directory containing rate sheets to fix')
    fix_parser.add_argument('--output', '-o', required=True, help='Output directory for cleaned files')

    args = parser.parse_args()

    if args.command == 'find-pdfs':
        return cmd_find_pdfs(args)
    elif args.command == 'parse-us-zones':
        return cmd_parse_us_zones(args)
    elif args.command == 'parse-ca-rates':
        return cmd_parse_ca_rates(args)
    elif args.command == 'generate':
        return cmd_generate(args)
    elif args.command == 'fix':
        return cmd_fix(args)
    else:
        parser.print_help()
        return 1


if __name__ == '__main__':
    sys.exit(main())
