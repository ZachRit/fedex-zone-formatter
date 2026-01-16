"""
Zone Rate Sheet Generator

Generates rate sheets by combining zone data with a rate sheet template.

Usage:
    python generate_rate_sheet.py \
      --input <ssl_postal_codes.xlsx> \
      --rate-sheet <rate_template.xlsx> \
      --country-name "United States" \
      --country-symbol "US" \
      --client-name "ClientName" \
      --carrier "FedEx" \
      --carrier-account "123456"
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def find_zone_file(postal_code: str, outputs_dir: str = 'output') -> Path | None:
    """
    Find the zone xlsx file whose filename range contains the postal code.

    Args:
        postal_code: The postal code to find (e.g., "01750")
        outputs_dir: Directory containing zone files

    Returns:
        Path to matching zone file, or None if no match found
    """
    outputs_path = Path(outputs_dir)

    # Normalize postal code to 5 digits
    postal_code = str(postal_code).zfill(5)
    postal_int = int(postal_code)

    # Iterate through all xlsx files in the output directory
    for zone_file in outputs_path.glob('*.xlsx'):
        # Skip temp files
        if zone_file.name.startswith('~$'):
            continue

        # Parse filename format: XXXXX-XXXXX.xlsx
        stem = zone_file.stem
        if '-' not in stem:
            continue

        try:
            start_str, end_str = stem.split('-')
            start_range = int(start_str)
            end_range = int(end_str)

            if start_range <= postal_int <= end_range:
                return zone_file
        except ValueError:
            continue

    return None


def load_zone_data(zone_file: Path, country_name: str, country_symbol: str) -> pd.DataFrame:
    """
    Load zone file and add Country Name, Country Symbol columns.

    Args:
        zone_file: Path to the zone xlsx file
        country_name: Country name to add (e.g., "United States")
        country_symbol: Country symbol to add (e.g., "US")

    Returns:
        DataFrame with columns: Country Name, Country Symbol, Zone, City,
                               Start Postal Code, End Postal Code
    """
    df = pd.read_excel(zone_file)

    # Add required columns
    df['Country Name'] = country_name
    df['Country Symbol'] = country_symbol
    df['City'] = ''  # Leave City empty

    # Format postal codes as 5-digit strings with leading zeros
    df['Start Postal Code'] = df['Start Postal Code'].apply(lambda x: str(int(x)).zfill(5) if pd.notna(x) else '')
    df['End Postal Code'] = df['End Postal Code'].apply(lambda x: str(int(x)).zfill(5) if pd.notna(x) else '')

    # Convert Zone to string, handle NaN (source column is 'Zone', output column is 'Zones')
    df['Zones'] = df['Zone'].apply(lambda x: str(int(x)) if pd.notna(x) else '')

    # Reorder columns to match rate sheet format
    df = df[['Country Name', 'Country Symbol', 'Zones', 'City', 'Start Postal Code', 'End Postal Code']]

    return df


def append_to_rate_sheet(rate_sheet_path: Path, zone_data: pd.DataFrame):
    """
    Load rate sheet, append zone data to Zone tab after row 3 headers.

    Args:
        rate_sheet_path: Path to the rate sheet template
        zone_data: DataFrame containing zone data to append

    Returns:
        Workbook object with appended data
    """
    wb = load_workbook(rate_sheet_path)

    if 'Zones' not in wb.sheetnames:
        raise ValueError(f"Rate sheet does not contain a 'Zones' tab")

    ws = wb['Zones']

    # Find the next empty row after row 3 (headers are on row 3)
    start_row = 4

    # Append data starting at row 4
    for idx, row in zone_data.iterrows():
        ws.cell(row=start_row + idx, column=1, value=row['Country Name'])
        ws.cell(row=start_row + idx, column=2, value=row['Country Symbol'])
        ws.cell(row=start_row + idx, column=3, value=row['Zones'])
        ws.cell(row=start_row + idx, column=4, value=row['City'])
        ws.cell(row=start_row + idx, column=5, value=row['Start Postal Code'])
        ws.cell(row=start_row + idx, column=6, value=row['End Postal Code'])

    return wb


def generate_output_filename(ssl: str, client_name: str, carrier: str, carrier_account: str) -> str:
    """
    Generate output filename: YYYYMMDD-{SSL}-{clientName}-{carrier}-{carrierAccount}.xlsx

    Args:
        ssl: SSL identifier
        client_name: Client name
        carrier: Carrier name (e.g., "FedEx")
        carrier_account: Carrier account number

    Returns:
        Formatted filename string
    """
    date_str = datetime.now().strftime('%Y%m%d')
    return f"{date_str}-{ssl}-{client_name}-{carrier}-{carrier_account}.xlsx"


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Generate rate sheets by combining zone data with a template'
    )
    parser.add_argument(
        '--input',
        required=True,
        help='Path to SSL/Postal Code xlsx file'
    )
    parser.add_argument(
        '--rate-sheet',
        required=True,
        help='Path to rate sheet template xlsx file'
    )
    parser.add_argument(
        '--country-name',
        required=True,
        help='Country name (e.g., "United States")'
    )
    parser.add_argument(
        '--country-symbol',
        required=True,
        help='Country symbol (e.g., "US")'
    )
    parser.add_argument(
        '--client-name',
        required=True,
        help='Client name for output filename'
    )
    parser.add_argument(
        '--carrier',
        required=True,
        help='Carrier name (e.g., "FedEx")'
    )
    parser.add_argument(
        '--carrier-account',
        required=True,
        help='Carrier account number'
    )

    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_args()

    # Validate input files exist
    input_path = Path(args.input)
    rate_sheet_path = Path(args.rate_sheet)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        sys.exit(1)

    if not rate_sheet_path.exists():
        print(f"Error: Rate sheet template not found: {rate_sheet_path}")
        sys.exit(1)

    # Read SSL/Postal Code file
    try:
        ssl_df = pd.read_excel(input_path)
    except Exception as e:
        print(f"Error reading input file: {e}")
        sys.exit(1)

    # Validate required columns
    required_cols = ['SSL', 'Postal Code']
    missing_cols = [col for col in required_cols if col not in ssl_df.columns]
    if missing_cols:
        print(f"Error: Input file missing required columns: {missing_cols}")
        sys.exit(1)

    # Validate rate sheet has Zone tab
    try:
        wb_test = load_workbook(rate_sheet_path)
        if 'Zones' not in wb_test.sheetnames:
            print(f"Error: Rate sheet does not contain a 'Zones' tab")
            sys.exit(1)
        wb_test.close()
    except Exception as e:
        print(f"Error reading rate sheet: {e}")
        sys.exit(1)

    # Group by SSL
    grouped = ssl_df.groupby('SSL')

    print(f"Processing {len(grouped)} SSL group(s)...")
    print()

    for ssl, group in grouped:
        print(f"Processing SSL: {ssl}")

        # Collect zone data for all postal codes in this SSL
        all_zone_data = []

        for _, row in group.iterrows():
            postal_code = str(row['Postal Code']).zfill(5)

            # Find matching zone file
            zone_file = find_zone_file(postal_code, 'output')

            if zone_file is None:
                print(f"  Warning: No zone file found for postal code {postal_code}, skipping")
                continue

            print(f"  Found zone file for {postal_code}: {zone_file.name}")

            # Load zone data
            zone_data = load_zone_data(zone_file, args.country_name, args.country_symbol)
            all_zone_data.append(zone_data)

        if not all_zone_data:
            print(f"  Warning: No zone data collected for SSL {ssl}, skipping")
            print()
            continue

        # Combine all zone data for this SSL
        combined_zone_data = pd.concat(all_zone_data, ignore_index=True)

        # Create output workbook from template
        wb = append_to_rate_sheet(rate_sheet_path, combined_zone_data)

        # Generate output filename and save
        output_filename = generate_output_filename(
            ssl, args.client_name, args.carrier, args.carrier_account
        )
        wb.save(output_filename)
        wb.close()

        print(f"  Output: {output_filename} ({len(combined_zone_data)} rows)")
        print()

    print("Complete!")


if __name__ == '__main__':
    main()
